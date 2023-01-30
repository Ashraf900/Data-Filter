import openpyxl

wb= openpyxl.load_workbook('1ST ROUND.xlsx')
sheet = wb.active


wb2= openpyxl.load_workbook('2ND ROUND.xlsx')
sheet2 = wb2.active

wb3= openpyxl.load_workbook('3RD MOP UP ROUND.xlsx')
sheet3 = wb3.active

wb4= openpyxl.load_workbook('4TH STRAY VACANCY ROUND .xlsx')
sheet4 = wb4.active

wb5= openpyxl.load_workbook('NEET PG 2022 ALL INDIA QUOTA SPECIAL STRAY VACANCY ROUND BRANCH-WISE CUTOFF LIST.xlsx')
sheet5 = wb5.active



data = []
head = ('Allotted Quota','Allotted Institute','Course','Alloted Category','Cutoff Rank')
R2_data=[]
R3_data=[]
R4_data=[]
R5_data=[]
f_data=[]
big_data =[]
count = 0 
for j in sheet.values:
       i = list(j)
      #  if ',' or ', ' or '  ' in i[1]:
      #     i[1]= i[1].replace(',', ' ')
      #     i[1]=i[1].replace(' ', ' ')
      #     i[1] = i[1].replace('  ', ' ')
      #     i[1] = i[1].replace('   ', ' ')

       data.append({head[0]:i[0],head[1]:i[1],head[2]:i[2],head[3]:i[3],'R1':i[4], 'R2':'-','Mop Up':'-','Stray Vacancy':'-','Special-stray vacancy':'-'})
       big_data.append({head[0]:i[0],head[1]:i[1],head[2]:i[2],head[3]:i[3],'R1':i[4], 'R2':'-','Mop Up':'-','Stray Vacancy':'-','Special-stray vacancy':'-'})
       count += 1



for i in sheet2.values:
  j = list(i)
  # if ',' or ', ' or '  ' in j[1]:
  #         j[1]= j[1].replace(',', ' ')
  #         j[1] = j[1].replace(' ', ' ')
  #         j[1] = j[1].replace('  ', ' ')
  #         j[1] = j[1].replace('   ', ' ')
  #         R2_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':j[4],'Mop Up':'-','Stray Vacancy':'-','Special-stray vacancy':'-'})
  # else:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  R2_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':j[4],'Mop Up':'-','Stray Vacancy':'-','Special-stray vacancy':'-'})





for items in R2_data:
  for f_items in big_data:
      if items['Allotted Institute'].lower()==f_items['Allotted Institute'].lower() and items['Course']==f_items['Course'] and items['Alloted Category']==f_items['Alloted Category']:
        f_items['R2']=items['R2']
        items['changed']='yes'
        
  if 'changed' not in items:
    big_data.append(items)


for i in sheet3.values:
  j = list(i)
  # if ',' or ', ' or '  ' in j[1]:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  #     R3_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':'-','Mop Up':j[4],'Stray Vacancy':'-','Special-stray vacancy':'-'})
  # else:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  R3_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':'-','Mop Up':j[4],'Stray Vacancy':'-','Special-stray vacancy':'-'})

for items in R3_data:
  for f_items in big_data:
      if items['Allotted Institute'].lower()==f_items['Allotted Institute'].lower() and items['Course']==f_items['Course'] and items['Alloted Category']==f_items['Alloted Category']:
        f_items['Mop Up']=items['Mop Up']
        items['changed']='yes'
        
  if 'changed' not in items:
    big_data.append(items)


for i in sheet4.values:
  j = list(i)
  # if ',' or ', ' or '  ' in j[1]:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  #     R4_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':'-','Mop Up':'-','Stray Vacancy':j[4],'Special-stray vacancy':'-'})
  # else:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  R4_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':'-','Mop Up':'-','Stray Vacancy':j[4],'Special-stray vacancy':'-'})

for items in R4_data:
  for f_items in big_data:
      if items['Allotted Institute'].lower()==f_items['Allotted Institute'].lower() and items['Course']==f_items['Course'] and items['Alloted Category']==f_items['Alloted Category']:
        f_items['Stray Vacancy']=items['Stray Vacancy']
        items['changed']='yes'
        
  if 'changed' not in items:
    big_data.append(items)



for i in sheet5.values:
  j = list(i)
  # if ',' or ', ' or '  ' in j[1]:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  #     R5_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':'-','Mop Up':'-','Stray Vacancy':'-','Special-stray vacancy':j[4]})
  # else:
  #     j[1]= j[1].replace(',', ' ')
  #     j[1] = j[1].replace(' ', ' ')
  #     j[1] = j[1].replace('  ', ' ')
  #     j[1] = j[1].replace('   ', ' ')
  R5_data.append({head[0]:j[0],head[1]:j[1],head[2]:j[2],head[3]:j[3],'R1':'-', 'R2':'-','Mop Up':'-','Stray Vacancy':'-','Special-stray vacancy':j[4]})

for items in R5_data:
  for f_items in big_data:
      if items['Allotted Institute'].lower()==f_items['Allotted Institute'].lower() and items['Course']==f_items['Course'] and items['Alloted Category']==f_items['Alloted Category']:
        f_items['Special-stray vacancy']=items['Special-stray vacancy']
        items['changed']='yes'
        
  if 'changed' not in items:
    big_data.append(items)

    
      


print(len(data))
print(len(R2_data))
print(len(R3_data))
print(len(R4_data))
print(len(big_data))
      

import pandas as pd



# convert into dataframe
df = pd.DataFrame(data=big_data)

#convert into excel
df.to_excel("Final.xlsx", index=False)
print("Dictionary converted into excel...")

